"""UPS monthly billing analysis pipeline.

Usage:
    python3 ups_pipeline.py 2026 04
    python3 ups_pipeline.py 2026 01

Reads UPS Billing Data File (250-column) CSVs from ~/Downloads,
filters to invoices in the requested month, parses charge codes,
builds a multi-tab Excel workbook with analysis, and saves it.
"""
import csv
import glob
import os
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DOWNLOADS = os.path.expanduser("/sessions/confident-amazing-noether/mnt/Downloads")
OUTPUTS = "/sessions/confident-amazing-noether/mnt/outputs"
CLAUDE_FILES = "/sessions/confident-amazing-noether/mnt/Claude Files"

# Field positions (1-indexed from UPS Billing Data File 2.1 spec)
F_ACCOUNT = 2
F_INVOICE_DATE = 5
F_INVOICE_NUMBER = 6
F_INVOICE_AMT = 11
F_CURRENCY = 10
F_PICKUP_DATE = 12
F_REF1 = 16
F_TRACKING = 21
F_BILLABLE_WT = 27
F_WT_UNIT = 28
F_ACTUAL_WT = 29
F_PKG_TYPE = 31
F_ZONE = 34
F_CHARGE_CAT = 44
F_CHARGE_CODE = 45
F_CHARGE_DESC = 46
F_PUBLISHED = 52
F_NET = 53
F_SHIPPER_NAME = 68
F_SHIPPER_CITY = 71
F_SHIPPER_POSTAL = 73
F_SHIPPER_COUNTRY = 74
F_RECEIVER_NAME = 76
F_RECEIVER_CITY = 79
F_RECEIVER_POSTAL = 81
F_RECEIVER_COUNTRY = 82

GBP = '£#,##0.00;[Red](£#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'
INT = '#,##0'

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
SECTION_FONT = Font(name="Arial", bold=True, size=12)
TOTAL_FILL = PatternFill("solid", start_color="FFE699")
TOTAL_FONT = Font(name="Arial", bold=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GLOSSARY = [
    ("Freight (FRT)",
     "The base shipping charge for a package — the published rate to move goods between origin and destination. "
     "Driven by service level (WW Expedited, Dom. Standard, etc.), zone, weight, and dimensions. "
     "The 'Net' figure is what you actually pay after corporate discount; the 'Published' figure is list price."),
    ("Fuel Surcharge (FSC)",
     "Variable percentage added to freight to cover diesel/jet fuel costs. UPS resets weekly based on fuel indexes. "
     "Typically 15–35% of freight; spikes correlate with fuel price moves. Track this MoM as a sanity check."),
    ("Service Surcharges",
     "Per-package add-ons that increase shipping cost: Surge Fee (PFC/PFR — peak/demand), Additional Handling (AHC/SAH — oversize), "
     "Remote Area / Extended Area (HIS/ESD/ESP/LDS), Residential (RES/REP), Pickup options (OSW/OFW/ASW/AFW), "
     "Returns Pickup Attempts (ART/RSO). These are real shipping costs driven by parcel characteristics."),
    ("Documentation",
     "Per-invoice paperwork and admin fees: Paper Commercial Invoice Surcharge (CIS — paid for paper customs docs; "
     "consider going paperless), Print Label (ALP), International Processing Fee (FIP), Duty & Tax Forwarding (F/D)."),
    ("Service Issues",
     "Penalties or unusual high-value charges worth investigating: Prohibited Item Fee (PIF — wrong content declaration), "
     "Customer Solution Service Fee (CGS — billing adjustment / dispute resolution). High-value lines here are good "
     "candidates for an internal review and potential UPS dispute."),
    ("Brokerage (BRK)",
     "Customs clearance services for international shipments — UPS acts as the broker. "
     "Includes disbursement fees, additional handling, entry preparation, document posting. "
     "Only applies to cross-border shipments."),
    ("VAT (Value Added Tax)",
     "Two flavours: (1) GOV/205 = VAT charged by the broker on imported goods value (often reclaimable for VAT-registered); "
     "(2) TAX/01 = 20% UK VAT on UPS service fees themselves (also reclaimable). Combined into the VAT bucket."),
    ("Duty / Tax",
     "Customs duty on imported goods. Calculated by HS commodity code and country of origin. "
     "Unlike VAT, duty is NOT reclaimable — it's a true cost of import."),
    ("Adjustment",
     "Post-invoice corrections — rebills, refunds, audit credits, supplemental fees applied after the original invoice."),
    ("Service Zone",
     "UPS routing zone (001 = same UK zone, 005-010 = international zones). Determines the base freight rate."),
    ("Tracking Number",
     "UPS package barcode (1Z prefix). One per physical parcel. The same tracking number may appear in multiple invoices "
     "(e.g. main shipping invoice + later customs/broker invoice + later VAT/duty rebill)."),
    ("Unique Shipments vs Shipment Lines",
     "'Unique Shipments' counts distinct tracking numbers (= physical parcels sent). "
     "'Shipment Lines' counts each tracking-invoice pair separately, so a parcel billed across two invoices counts twice. "
     "When the two diverge, it indicates adjustment or supplementary invoices were issued."),
]


# ACC sub-categorisation. Anything not listed defaults to "Service Surcharges".
ACC_DOCUMENTATION = {"CIS", "ALP", "FIP", "F/D"}      # paperwork / admin
ACC_SERVICE_ISSUES = {"PIF", "CGS"}                    # penalties / unusual high-value


def categorize(cat, code, desc):
    if cat == "FRT":
        return ("Freight", desc or "Freight")
    if cat == "FSC":
        return ("Fuel Surcharge", "Fuel Surcharge")
    if cat == "ACC":
        if code in ACC_DOCUMENTATION:
            return ("Documentation", desc or f"Documentation {code}")
        if code in ACC_SERVICE_ISSUES:
            return ("Service Issues", desc or f"Service Issue {code}")
        return ("Service Surcharges", desc or f"Surcharge {code}")
    if cat == "BRK":
        return ("Brokerage", desc or f"Brokerage {code}")
    if cat == "GOV":
        if code == "205" or "Value Added Tax" in (desc or ""):
            return ("VAT", "Value Added Tax")
        return ("Duty/Tax", desc or f"Govt {code}")
    if cat == "EXM":
        return ("Exempt (info)", desc or "Exempt")
    if cat == "INF":
        return ("Info", desc or "Info")
    if cat == "ADJ":
        return ("Adjustment", desc or f"Adjustment {code}")
    if cat == "TAX":
        return ("VAT", desc or "Tax")
    if cat == "MSC":
        return ("Misc", desc or f"Misc {code}")
    if cat == "":
        return ("Other/Blank", "")
    return ("Other", f"{cat} {code}".strip())


def parse_file(path):
    rows = []
    with open(path, encoding="utf-8") as f:
        reader = csv.reader(f)
        for r in reader:
            if len(r) < F_NET:
                continue
            cat = r[F_CHARGE_CAT - 1]
            code = r[F_CHARGE_CODE - 1]
            desc = r[F_CHARGE_DESC - 1]
            try:
                published = float(r[F_PUBLISHED - 1] or 0)
                net = float(r[F_NET - 1] or 0)
            except ValueError:
                published, net = 0.0, 0.0
            bucket, friendly = categorize(cat, code, desc)
            rows.append({
                "Account": r[F_ACCOUNT - 1],
                "Invoice Date": r[F_INVOICE_DATE - 1],
                "Invoice Number": r[F_INVOICE_NUMBER - 1],
                "Invoice Total": r[F_INVOICE_AMT - 1],
                "Pickup Date": r[F_PICKUP_DATE - 1],
                "Tracking Number": r[F_TRACKING - 1],
                "Sales Order Ref": r[F_REF1 - 1],
                "Service Zone": r[F_ZONE - 1],
                "Billable Wt": r[F_BILLABLE_WT - 1],
                "Wt Unit": r[F_WT_UNIT - 1],
                "Actual Wt": r[F_ACTUAL_WT - 1],
                "Pkg Type": r[F_PKG_TYPE - 1],
                "Charge Cat": cat, "Charge Code": code, "Charge Desc": desc,
                "Bucket": bucket, "Friendly Name": friendly,
                "Published": published, "Net": net,
                "Currency": r[F_CURRENCY - 1],
                "Origin Name": r[F_SHIPPER_NAME - 1],
                "Origin City": r[F_SHIPPER_CITY - 1],
                "Origin Postal": r[F_SHIPPER_POSTAL - 1],
                "Origin Country": r[F_SHIPPER_COUNTRY - 1],
                "Dest Name": r[F_RECEIVER_NAME - 1],
                "Dest City": r[F_RECEIVER_CITY - 1],
                "Dest Postal": r[F_RECEIVER_POSTAL - 1],
                "Dest Country": r[F_RECEIVER_COUNTRY - 1],
                "Source File": os.path.basename(path),
            })
    return rows


def collect_month(year, month):
    """Find all Invoice CSVs in Downloads with invoice dates in the given year-month."""
    files = sorted(set(glob.glob(f"{DOWNLOADS}/Invoice_*.csv")))
    matched = []
    for path in files:
        if "(1)" in path:  # skip duplicates
            continue
        with open(path, encoding="utf-8") as f:
            r = next(csv.reader(f), None)
        if not r or len(r) < F_INVOICE_DATE:
            continue
        date = r[F_INVOICE_DATE - 1]
        if date.startswith(f"{year:04d}-{month:02d}-"):
            matched.append(path)
    return matched


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        col = col_cells[0].column_letter
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col].width = max(min_w, min(max_w, length + 2))


def short_dest(s):
    """Build 'Customer / City, CC' summary string for a shipment."""
    name = (s["Dest Name"] or "").strip().title()[:30]
    city = (s["Dest City"] or "").strip().title()
    ctry = (s["Dest Country"] or "").strip().upper()
    if not (name or city):
        return ""
    parts = []
    if name:
        parts.append(name)
    if city:
        parts.append(city)
    if ctry:
        parts.append(ctry)
    return " / ".join(parts) if name else f'{city}, {ctry}' if city else ctry


def short_origin(s):
    name = (s["Origin Name"] or "").strip().title()[:30]
    city = (s["Origin City"] or "").strip().title()
    ctry = (s["Origin Country"] or "").strip().upper()
    parts = []
    if name:
        parts.append(name)
    if city:
        parts.append(city)
    if ctry:
        parts.append(ctry)
    return " / ".join(parts)


def build_workbook(rows, year, month, out_path):
    invoice_set = {r["Invoice Number"] for r in rows}
    tracking_set = {r["Tracking Number"] for r in rows
                    if r["Tracking Number"] and r["Tracking Number"].startswith("1Z")}
    n_invoices = len(invoice_set)
    n_shipments = len(tracking_set)
    month_name = ["", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"][month]

    wb = Workbook()

    # ===== Tab 1: Summary =====
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"UPS Billing Analysis — {month_name} {year}"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E78")
    ws["A2"] = f"Account 6V41A5  •  {n_invoices} invoices  •  {n_shipments} unique trackings  •  Source: UPS Billing Centre 250-col CSV"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")

    # Pre-compute additional metrics
    inv_tracking_pairs = set()
    for r in rows:
        if r["Tracking Number"] and r["Tracking Number"].startswith("1Z"):
            inv_tracking_pairs.add((r["Invoice Number"], r["Tracking Number"]))
    n_shipment_lines = len(inv_tracking_pairs)

    ws["A4"] = "Total Net Charges"
    ws["A5"] = "Unique Shipments (parcels)"
    ws["A6"] = "Shipment Lines (billed events)"
    ws["A7"] = "Number of Invoices"
    ws["A8"] = "Avg cost / parcel"
    ws["B4"] = '=SUMIFS(Charges!S:S,Charges!M:M,"<>EXM",Charges!M:M,"<>INF")'
    ws["B5"] = n_shipments
    ws["B6"] = n_shipment_lines
    ws["B7"] = n_invoices
    ws["B8"] = "=B4/B5"
    ws["B4"].number_format = GBP
    ws["B5"].number_format = INT
    ws["B6"].number_format = INT
    ws["B7"].number_format = INT
    ws["B8"].number_format = GBP
    for r in (4, 5, 6, 7, 8):
        ws.cell(row=r, column=1).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=2).font = Font(name="Arial", bold=True)

    ws["A9"] = "Charges by category"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL
    ws.merge_cells("A9:D9")

    ws.append([])
    ws.append(["Category", "Net (£)", "% of total", "# of charges"])
    style_header(ws, ws.max_row, 4)

    buckets_order = ["Freight", "Fuel Surcharge", "Service Surcharges", "Documentation",
                     "Service Issues", "Brokerage", "VAT", "Duty/Tax",
                     "Adjustment", "Misc", "Other", "Other/Blank"]
    start = ws.max_row + 1
    for b in buckets_order:
        ws.append([
            b,
            f'=SUMIF(Charges!P:P,"{b}",Charges!S:S)',
            f"=IFERROR(B{ws.max_row+1}/$B$4,0)",
            f'=COUNTIF(Charges!P:P,"{b}")',
        ])
    end = ws.max_row
    ws.append(["Total", f"=SUM(B{start}:B{end})", f"=SUM(C{start}:C{end})", f"=SUM(D{start}:D{end})"])
    total_row = ws.max_row
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=c).font = TOTAL_FONT
    for r in range(start, total_row + 1):
        ws.cell(row=r, column=2).number_format = GBP
        ws.cell(row=r, column=3).number_format = PCT
        ws.cell(row=r, column=4).number_format = INT

    ws.append([])
    ws.append(["Fuel Surcharge as % of Freight",
               '=SUMIF(Charges!P:P,"Fuel Surcharge",Charges!S:S)/SUMIF(Charges!P:P,"Freight",Charges!S:S)'])
    ws.cell(row=ws.max_row, column=1).font = TOTAL_FONT
    ws.cell(row=ws.max_row, column=2).number_format = PCT
    ws.cell(row=ws.max_row, column=2).font = TOTAL_FONT

    autofit(ws, min_w=14, max_w=48)
    ws.column_dimensions["A"].width = 32

    # ===== Tab 2: Glossary =====
    ws_g = wb.create_sheet("Glossary")
    ws_g["A1"] = "What each charge category means"
    ws_g["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    ws_g.append([])
    ws_g.append(["Category", "Definition"])
    style_header(ws_g, ws_g.max_row, 2)
    for term, definition in GLOSSARY:
        ws_g.append([term, definition])
        ws_g.cell(row=ws_g.max_row, column=1).font = Font(name="Arial", bold=True)
        ws_g.cell(row=ws_g.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_g.column_dimensions["A"].width = 26
    ws_g.column_dimensions["B"].width = 100
    for r in range(4, ws_g.max_row + 1):
        ws_g.row_dimensions[r].height = 60

    # ===== Tab 3: By Invoice =====
    ws_inv = wb.create_sheet("By Invoice")
    inv_buckets = ["Freight", "Fuel Surcharge", "Service Surcharges", "Documentation",
                   "Service Issues", "Brokerage", "VAT", "Duty/Tax",
                   "Adjustment", "Misc", "Other"]
    inv_headers = ["Invoice Number", "Invoice Date", "# Shipments"] + inv_buckets + ["Net Total"]
    ws_inv.append(inv_headers)
    style_header(ws_inv, 1, len(inv_headers))

    invoices = {}
    for r in rows:
        inv = r["Invoice Number"]
        if inv not in invoices:
            invoices[inv] = {"date": r["Invoice Date"], "shipments": set()}
        if r["Tracking Number"] and r["Tracking Number"].startswith("1Z"):
            invoices[inv]["shipments"].add(r["Tracking Number"])
    invoice_list = sorted(invoices.items(), key=lambda x: x[1]["date"])

    # Bucket columns are D..N (11 buckets), Net Total in O
    n_buckets = len(inv_buckets)
    last_bucket_col = chr(ord("D") + n_buckets - 1)  # N
    net_total_col = chr(ord("D") + n_buckets)        # O
    for inv, meta in invoice_list:
        row_data = [inv, meta["date"], len(meta["shipments"])]
        for b in inv_buckets:
            row_data.append(f'=SUMIFS(Charges!S:S,Charges!C:C,"{inv}",Charges!P:P,"{b}")')
        row_data.append(f"=SUM(D{ws_inv.max_row+1}:{last_bucket_col}{ws_inv.max_row+1})")
        ws_inv.append(row_data)
    last_data = ws_inv.max_row
    total_cells = ["TOTAL", "", f"=SUM(C2:C{last_data})"]
    for i in range(n_buckets):
        col_letter = chr(ord("D") + i)
        total_cells.append(f"=SUM({col_letter}2:{col_letter}{last_data})")
    total_cells.append(f"=SUM({net_total_col}2:{net_total_col}{last_data})")
    ws_inv.append(total_cells)
    trow = ws_inv.max_row
    for c in range(1, len(inv_headers) + 1):
        ws_inv.cell(row=trow, column=c).fill = TOTAL_FILL
        ws_inv.cell(row=trow, column=c).font = TOTAL_FONT
    for r in range(2, trow + 1):
        for c in range(4, len(inv_headers) + 1):
            ws_inv.cell(row=r, column=c).number_format = GBP
        ws_inv.cell(row=r, column=3).number_format = INT
    autofit(ws_inv, min_w=12, max_w=22)

    # ===== Tab 4: Shipments =====
    ws_sh = wb.create_sheet("Shipments")
    sh_buckets = ["Freight", "Fuel Surcharge", "Service Surcharges", "Documentation",
                  "Service Issues", "Brokerage", "VAT", "Duty/Tax",
                  "Adjustment", "Misc", "Other"]
    sh_headers = [
        "Invoice Number", "Invoice Date", "Pickup Date", "Tracking Number",
        "Sales Order Ref", "Service", "Service Zone",
        "Origin Name", "Origin City", "Origin Country",
        "Dest Name", "Dest City", "Dest Country",
        "Billable Wt", "Wt Unit",
    ] + sh_buckets + ["Net Total", "Fuel % of Freight"]
    ws_sh.append(sh_headers)
    style_header(ws_sh, 1, len(sh_headers))

    ship = {}
    for r in rows:
        tk = r["Tracking Number"]
        if not tk or not tk.startswith("1Z"):
            continue
        key = (r["Invoice Number"], tk)
        if key not in ship:
            ship[key] = {
                "Invoice Number": r["Invoice Number"],
                "Invoice Date": r["Invoice Date"],
                "Pickup Date": r["Pickup Date"],
                "Tracking Number": tk,
                "Sales Order Ref": r["Sales Order Ref"],
                "Service": "",
                "Service Zone": r["Service Zone"],
                "Origin Name": r["Origin Name"],
                "Origin City": r["Origin City"],
                "Origin Country": r["Origin Country"],
                "Dest Name": r["Dest Name"],
                "Dest City": r["Dest City"],
                "Dest Country": r["Dest Country"],
                "Billable Wt": r["Billable Wt"],
                "Wt Unit": r["Wt Unit"],
                "buckets": defaultdict(float),
            }
        for fld in ("Origin Name", "Origin City", "Origin Country",
                    "Dest Name", "Dest City", "Dest Country"):
            if not ship[key][fld] and r[fld]:
                ship[key][fld] = r[fld]
        if r["Charge Cat"] == "FRT" and r["Charge Desc"]:
            ship[key]["Service"] = r["Charge Desc"]
        if r["Charge Cat"] not in ("EXM", "INF"):
            ship[key]["buckets"][r["Bucket"]] += r["Net"]

    # Bucket columns start at P (col 16). Use get_column_letter for AA/AB beyond Z.
    from openpyxl.utils import get_column_letter
    n_b = len(sh_buckets)
    first_bucket_col = get_column_letter(16)
    last_bucket_col = get_column_letter(16 + n_b - 1)
    net_total_col = get_column_letter(16 + n_b)
    fuel_col = get_column_letter(17)
    freight_col = get_column_letter(16)
    for key, s in sorted(ship.items(), key=lambda x: (x[1]["Invoice Date"], x[1]["Tracking Number"])):
        row = [
            s["Invoice Number"], s["Invoice Date"], s["Pickup Date"], s["Tracking Number"],
            s["Sales Order Ref"], s["Service"], s["Service Zone"],
            s["Origin Name"], s["Origin City"], s["Origin Country"],
            s["Dest Name"], s["Dest City"], s["Dest Country"],
            float(s["Billable Wt"] or 0), s["Wt Unit"],
        ]
        for b in sh_buckets:
            row.append(s["buckets"].get(b, 0))
        n = ws_sh.max_row + 1
        row.append(f"=SUM({first_bucket_col}{n}:{last_bucket_col}{n})")
        row.append(f"=IFERROR({fuel_col}{n}/{freight_col}{n},0)")
        ws_sh.append(row)
    last = ws_sh.max_row
    bucket_first_col_idx = 16  # P
    bucket_last_col_idx = bucket_first_col_idx + n_b - 1  # Z
    net_total_col_idx = bucket_last_col_idx + 1
    fuel_pct_col_idx = bucket_last_col_idx + 2
    for r in range(2, last + 1):
        for col in range(bucket_first_col_idx, net_total_col_idx + 1):
            ws_sh.cell(row=r, column=col).number_format = GBP
        ws_sh.cell(row=r, column=fuel_pct_col_idx).number_format = PCT
    total = ["TOTAL"] + [""] * 14
    for i in range(n_b + 1):  # buckets + Net Total
        col_letter = get_column_letter(16 + i)
        total.append(f"=SUM({col_letter}2:{col_letter}{last})")
    total.append("")
    ws_sh.append(total)
    trow = ws_sh.max_row
    for c in range(1, len(sh_headers) + 1):
        ws_sh.cell(row=trow, column=c).fill = TOTAL_FILL
        ws_sh.cell(row=trow, column=c).font = TOTAL_FONT
    for c in range(bucket_first_col_idx, net_total_col_idx + 1):
        ws_sh.cell(row=trow, column=c).number_format = GBP
    autofit(ws_sh, min_w=10, max_w=24)
    ws_sh.freeze_panes = "E2"

    # ===== Tab 5: Top Surcharges =====
    ws_top = wb.create_sheet("Top Surcharges")
    ws_top.append(["Charge Description", "Bucket", "Code", "# of charges", "Total Net (£)"])
    style_header(ws_top, 1, 5)
    acc_totals = defaultdict(lambda: {"count": 0, "net": 0.0, "code": "", "bucket": ""})
    track_buckets = {"Service Surcharges", "Documentation", "Service Issues",
                     "Fuel Surcharge", "Brokerage", "Duty/Tax", "Other"}
    for r in rows:
        if r["Bucket"] not in track_buckets:
            continue
        if r["Charge Cat"] in ("EXM", "INF"):
            continue
        if abs(r["Net"]) < 0.01:
            continue
        key = r["Friendly Name"] or f'{r["Charge Cat"]} {r["Charge Code"]}'
        acc_totals[key]["count"] += 1
        acc_totals[key]["net"] += r["Net"]
        acc_totals[key]["code"] = r["Charge Code"]
        acc_totals[key]["bucket"] = r["Bucket"]
    for desc, d in sorted(acc_totals.items(), key=lambda x: -x[1]["net"]):
        ws_top.append([desc, d["bucket"], d["code"], d["count"], d["net"]])
    last = ws_top.max_row
    for r in range(2, last + 1):
        ws_top.cell(row=r, column=4).number_format = INT
        ws_top.cell(row=r, column=5).number_format = GBP
    autofit(ws_top, min_w=12, max_w=40)

    # ===== Tab 5b: Charge Detail (every line per shipment, in tracking order) =====
    ws_cd = wb.create_sheet("Charge Detail")
    cd_headers = [
        "Invoice Number", "Invoice Date", "Tracking Number", "Sales Order Ref",
        "Service Zone", "Charge Cat", "Charge Code", "Charge Description",
        "Bucket", "Published", "Net",
    ]
    ws_cd.append(cd_headers)
    style_header(ws_cd, 1, len(cd_headers))
    # Sort by tracking, then by category to keep all lines for a tracking grouped
    cat_order = {"FRT": 0, "FSC": 1, "ACC": 2, "BRK": 3, "GOV": 4, "TAX": 5,
                 "ADJ": 6, "MSC": 7, "EXM": 8, "INF": 9}
    sorted_rows = sorted(
        [r for r in rows if r["Tracking Number"] and r["Tracking Number"].startswith("1Z")],
        key=lambda x: (x["Invoice Date"], x["Invoice Number"], x["Tracking Number"],
                       cat_order.get(x["Charge Cat"], 99))
    )
    last_inv_track = None
    for r in sorted_rows:
        ws_cd.append([
            r["Invoice Number"], r["Invoice Date"], r["Tracking Number"], r["Sales Order Ref"],
            r["Service Zone"], r["Charge Cat"], r["Charge Code"], r["Charge Desc"],
            r["Bucket"], r["Published"], r["Net"],
        ])
        # Lightly shade alternating tracking groups so it's easy to read
        cur = (r["Invoice Number"], r["Tracking Number"])
        if cur != last_inv_track:
            last_inv_track = cur
    for r in range(2, ws_cd.max_row + 1):
        ws_cd.cell(row=r, column=10).number_format = GBP
        ws_cd.cell(row=r, column=11).number_format = GBP
    autofit(ws_cd, min_w=10, max_w=40)
    ws_cd.freeze_panes = "D2"

    # ===== Tab 6: Charges (raw long) =====
    ws_ch = wb.create_sheet("Charges")
    headers = [
        "Account", "Invoice Date", "Invoice Number", "Invoice Total",
        "Pickup Date", "Tracking Number", "Sales Order Ref",
        "Service Zone", "Billable Wt", "Wt Unit", "Actual Wt", "Pkg Type",
        "Charge Cat", "Charge Code", "Charge Desc", "Bucket", "Friendly Name",
        "Published", "Net", "Currency",
        "Origin Name", "Origin City", "Origin Postal", "Origin Country",
        "Dest Name", "Dest City", "Dest Postal", "Dest Country",
        "Source File",
    ]
    ws_ch.append(headers)
    style_header(ws_ch, 1, len(headers))
    for r in rows:
        ws_ch.append([r.get(h, "") for h in headers])
    last = ws_ch.max_row
    for r in range(2, last + 1):
        ws_ch.cell(row=r, column=4).number_format = GBP
        ws_ch.cell(row=r, column=18).number_format = GBP
        ws_ch.cell(row=r, column=19).number_format = GBP
    autofit(ws_ch, min_w=10, max_w=24)
    ws_ch.freeze_panes = "C2"

    wb.save(out_path)
    return n_invoices, n_shipments



def main():
    if len(sys.argv) < 3:
        print("Usage: ups_pipeline.py YYYY MM")
        sys.exit(1)
    year = int(sys.argv[1])
    month = int(sys.argv[2])
    files = collect_month(year, month)
    print(f"Found {len(files)} invoice files for {year}-{month:02d}")
    if not files:
        sys.exit(1)
    rows = []
    for path in files:
        rows.extend(parse_file(path))
    month_name = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][month]
    out_xlsx = os.path.join(OUTPUTS, f"UPS_{month_name}_{year}_Analysis.xlsx")
    n_inv, n_ship = build_workbook(rows, year, month, out_xlsx)
    invoice_actual = defaultdict(float)
    invoice_field = {}
    for r in rows:
        if r["Charge Cat"] in ("EXM", "INF"):
            continue
        invoice_actual[r["Invoice Number"]] += r["Net"]
        invoice_field[r["Invoice Number"]] = r["Invoice Total"]
    grand_sum = sum(invoice_actual.values())
    grand_field = sum(float(v or 0) for v in invoice_field.values() if v)
    print(f"GRAND TOTAL: sum_net={grand_sum:.2f}, field_total={grand_field:.2f}")
    print(f"Wrote: {out_xlsx}")
    print(f"  Invoices: {n_inv}, Shipments: {n_ship}")


if __name__ == "__main__":
    main()
